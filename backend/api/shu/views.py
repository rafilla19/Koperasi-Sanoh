# shu/views.py
from decimal import Decimal, InvalidOperation
from datetime import datetime

from django.db.models import Q, Sum, F
from django.db.models.functions import TruncWeek
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response

from api.models import Members, MemberBankAccounts
from api.saving.models import MemberSavingObligations, SavingTransactions, SavingWallets
from api.models import IncomeExpenseCategories, IncomeExpenses  # noqa: F401 — IncomeExpenseCategories used for TYPE constants
from .models import (
    AccountingPeriods, MasterConfiguration,
    ShuPeriods, ShuMemberDistributions, ShuMemberDistributionsMonthly,
    ShuResults, ShuMemberBases, ShuComponentAllocation,
)
from django.db import transaction, IntegrityError, connection
from .forecasting import forecast_shu
from .serializers import (
    AdminAnnualJasaModalSerializer,
    AdminShuDistributionSerializer,
    IncomeExpenseCategorySerializer,
    IncomeExpenseOutcomeCreateSerializer,
    IncomeExpenseOutcomeSerializer,
    MasterConfigurationSerializer,
    MemberShuDistributionSerializer,
    ShuDistributionUpdateSerializer,
    ShuPeriodCreateSerializer,
    ShuPeriodsSerializer,
    ShuResultsSerializer,
)

TEMP_MEMBER_ID = 5


def _get_bank_map(member_ids):
    """Fetch bank info keyed by member_id using raw SQL (MemberBankAccount has no bank FK)."""
    if not member_ids:
        return {}
    placeholders = ','.join(['%s'] * len(member_ids))
    with connection.cursor() as cursor:
        cursor.execute(
            f"SELECT mba.member_id, b.bank_name, b.bank_code, mba.account_number, mba.account_holder_name "
            f"FROM member_bank_accounts mba "
            f"JOIN banks b ON b.id = mba.bank_id "
            f"WHERE mba.member_id IN ({placeholders})",
            list(member_ids),
        )
        rows = cursor.fetchall()
    bank_map = {}
    for row in rows:
        mid = row[0]
        if mid not in bank_map:
            bank_map[mid] = {
                'bank_name': row[1],
                'bank_code': row[2],
                'account_number': row[3],
                'account_holder_name': row[4],
            }
    return bank_map


# ── MASTER CONFIGURATION ─────────────────────────────────────────

@api_view(['GET', 'POST'])
def admin_master_configurations(request):
    """
    GET  → list all master configuration items (component_name + percentage).
    POST → create new item. Validates total percentage does not exceed 100.
    """
    if request.method == 'POST':
        serializer = MasterConfigurationSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)

        new_pct = serializer.validated_data['percentage']
        current_total = MasterConfiguration.objects.filter(
            deleted_at__isnull=True
        ).aggregate(total=Sum('percentage'))['total'] or Decimal('0')

        if current_total + new_pct > Decimal('100'):
            return Response(
                {'error': f'Total persentase melebihi 100%. Sisa: {100 - float(current_total)}%'},
                status=400,
            )

        item = serializer.save()
        total_after = float(current_total + new_pct)
        return Response({
            **MasterConfigurationSerializer(item).data,
            'total_percentage': total_after,
        }, status=201)

    items = MasterConfiguration.objects.filter(deleted_at__isnull=True).order_by('id')
    total = items.aggregate(total=Sum('percentage'))['total'] or Decimal('0')
    return Response({
        'total_percentage': float(total),
        'results': MasterConfigurationSerializer(items, many=True).data,
    })


@api_view(['PATCH', 'DELETE'])
def admin_master_configuration_detail(request, pk):
    """
    PATCH  → update percentage (and optionally name) of one item.
    DELETE → remove item.
    Validates total percentage does not exceed 100 on update.
    """
    try:
        item = MasterConfiguration.objects.get(pk=pk)
    except MasterConfiguration.DoesNotExist:
        return Response({'error': 'Item tidak ditemukan'}, status=404)

    if request.method == 'DELETE':
        item.delete()
        return Response(status=204)

    serializer = MasterConfigurationSerializer(item, data=request.data, partial=True)
    if not serializer.is_valid():
        return Response(serializer.errors, status=400)

    if 'percentage' in serializer.validated_data:
        new_pct = serializer.validated_data['percentage']
        other_total = MasterConfiguration.objects.filter(
            deleted_at__isnull=True
        ).exclude(pk=pk).aggregate(total=Sum('percentage'))['total'] or Decimal('0')

        if other_total + new_pct > Decimal('100'):
            return Response(
                {'error': f'Total persentase melebihi 100%. Sisa untuk item ini: {100 - float(other_total)}%'},
                status=400,
            )

    serializer.save()
    total = MasterConfiguration.objects.filter(
        deleted_at__isnull=True
    ).aggregate(total=Sum('percentage'))['total'] or Decimal('0')
    return Response({
        **MasterConfigurationSerializer(item).data,
        'total_percentage': float(total),
    })


# ── MEMBER ───────────────────────────────────────────────────────

@api_view(['GET'])
def my_shu_analytics(request):
    """
    SHU Analytics data for logged-in member.
    Returns unpaid total SHU, monthly chart data, and growth percentage.
    """
    member_id = request.query_params.get('member_id')
    if not member_id:
        if request.user.is_authenticated and hasattr(request.user, 'member'):
            member_id = request.user.member.id
        else:
            member_id = TEMP_MEMBER_ID

    # 1. Total SHU (where distributed_status = false)
    total_shu_qs = ShuMemberDistributionsMonthly.objects.filter(
        member_id=member_id,
        distributed_status=False
    ).aggregate(total_shu=Sum('total_shu'))
    total_shu = total_shu_qs['total_shu'] or Decimal('0')

    # 2. Data Chart Per Bulan — use FK period → ShuResults for reliable month
    monthly_data = (
        ShuMemberDistributionsMonthly.objects.filter(
            member_id=member_id,
            period__deleted_at__isnull=True,
            period__period_month__gte=1,
            period__period_month__lte=12,
        )
        .values(
            p_year=F('period__period_year'),
            p_month=F('period__period_month'),
        )
        .annotate(total_shu=Sum('total_shu'))
        .order_by('p_year', 'p_month')
    )

    chart_data = []
    for row in monthly_data:
        chart_data.append({
            'month': f"{row['p_year']}-{row['p_month']:02d}",
            'total_shu': float(row['total_shu'] or 0)
        })

    # 3. Growth Percentage based on the last 2 months
    growth_percentage = 0.0
    previous_total = 0.0
    if len(chart_data) >= 2:
        c_total = chart_data[-1]['total_shu']
        p_total = chart_data[-2]['total_shu']
        previous_total = p_total
        if p_total > 0:
            growth_percentage = round(((c_total - p_total) / p_total) * 100, 1)

    # 4. ML Forecast
    forecast_result = None
    try:
        forecast_result = forecast_shu(chart_data)
    except Exception:
        import logging
        logging.getLogger(__name__).exception("SHU forecast failed")

    # 5. Current-year monthly SHU breakdown
    current_year = timezone.now().year
    current_year_monthly = (
        ShuMemberDistributionsMonthly.objects.filter(
            member_id=member_id,
            period__period_year=current_year,
            period__period_month__gte=1,
            period__period_month__lte=12,
            period__deleted_at__isnull=True,
        )
        .values(p_month=F('period__period_month'))
        .annotate(
            total_shu=Sum('total_shu'),
            total_savings=Sum('total_savings'),
        )
        .order_by('p_month')
    )
    current_year_data = []
    current_year_total = Decimal('0')
    for row in current_year_monthly:
        shu_val = row['total_shu'] or Decimal('0')
        current_year_total += shu_val
        current_year_data.append({
            'month': row['p_month'],
            'total_shu': float(shu_val),
            'total_savings': float(row['total_savings'] or 0),
        })

    # 6. Yearly SHU history from shu_member_distributions
    yearly_history = list(
        ShuMemberDistributions.objects.filter(
            member_id=member_id,
            period_year__isnull=False,
        )
        .values('period_year')
        .annotate(
            total_shu=Sum('total_shu'),
            total_savings=Sum('total_savings'),
        )
        .order_by('-period_year')
    )
    yearly_data = []
    for row in yearly_history:
        yearly_data.append({
            'year': row['period_year'],
            'total_shu': float(row['total_shu'] or 0),
            'total_savings': float(row['total_savings'] or 0),
        })

    return Response({
        'total_shu': float(total_shu),
        'chart_data': chart_data,
        'growth_percentage': growth_percentage,
        'previous_total': previous_total,
        'forecast': forecast_result,
        'current_year': {
            'year': current_year,
            'months': current_year_data,
            'total_shu': float(current_year_total),
        },
        'yearly_history': yearly_data,
    })


@api_view(['GET'])
def my_shu_distributions(request):
    """Daftar SHU yang diterima member yang sedang login."""
    distributions = ShuMemberDistributions.objects.filter(
        member_id=TEMP_MEMBER_ID,
    ).order_by('-period_year')
    return Response(MemberShuDistributionSerializer(distributions, many=True).data)


@api_view(['GET'])
def my_shu_detail(request, pk):
    """Detail SHU untuk satu periode tertentu."""
    try:
        dist = ShuMemberDistributions.objects.get(pk=pk, member_id=TEMP_MEMBER_ID)
    except ShuMemberDistributions.DoesNotExist:
        return Response({'error': 'Data SHU tidak ditemukan'}, status=404)
    return Response(MemberShuDistributionSerializer(dist).data)


# ── ADMIN: PERIOD ────────────────────────────────────────────────

@api_view(['GET'])
def admin_shu_periods(request):
    """Daftar semua periode SHU."""
    periods = ShuPeriods.objects.all().order_by('-year')
    return Response(ShuPeriodsSerializer(periods, many=True).data)


@api_view(['POST'])
def admin_shu_period_create(request):
    """
    Buat periode SHU baru.
    Body: { year, total_profit, total_savings_weight, total_transaction_weight,
            member_services_weight, reserve_fund_weight, social_fund_weight,
            education_fund_weight, management_weight, notes? }
    """
    serializer = ShuPeriodCreateSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=400)

    data = serializer.validated_data
    if ShuPeriods.objects.filter(year=data['year']).exists():
        return Response({'error': f'Periode SHU tahun {data["year"]} sudah ada'}, status=400)

    period = ShuPeriods.objects.create(**data)
    return Response(ShuPeriodsSerializer(period).data, status=201)


@api_view(['GET', 'PATCH'])
def admin_shu_period_detail(request, pk):
    """
    GET  → detail periode SHU.
    PATCH → update field periode (total_profit, bobot, status, notes).
    """
    try:
        period = ShuPeriods.objects.get(pk=pk)
    except ShuPeriods.DoesNotExist:
        return Response({'error': 'Periode tidak ditemukan'}, status=404)

    if request.method == 'GET':
        return Response(ShuPeriodsSerializer(period).data)

    allowed_fields = {
        'total_profit', 'total_savings_weight', 'total_transaction_weight',
        'member_services_weight', 'reserve_fund_weight', 'social_fund_weight',
        'education_fund_weight', 'management_weight', 'status', 'notes',
    }
    for field, value in request.data.items():
        if field in allowed_fields:
            setattr(period, field, value)
    period.save()
    return Response(ShuPeriodsSerializer(period).data)


# ── ADMIN: CALCULATE & DISTRIBUTE ───────────────────────────────

@api_view(['POST'])
def admin_shu_calculate(request, pk):
    """
    Hitung SHU per member untuk periode ini berdasarkan:
    - Saldo simpanan rata-rata (dari saving_wallets)
    - Total transaksi (dari saving_transactions, tipe kredit)
    Bobot member_services_weight dibagi dua komponen tsb sesuai konfigurasi period.
    """
    try:
        period = ShuPeriods.objects.get(pk=pk)
    except ShuPeriods.DoesNotExist:
        return Response({'error': 'Periode tidak ditemukan'}, status=404)

    if period.status not in ('draft', 'calculated'):
        return Response({'error': 'Hanya periode berstatus draft atau calculated yang bisa dihitung ulang'}, status=400)

    members = Members.objects.filter(deleted_at__isnull=True)

    # Total saldo simpanan semua member
    total_savings = SavingWallets.objects.filter(
        deleted_at__isnull=True
    ).aggregate(total=Sum('balance'))['total'] or Decimal('0')

    # Total transaksi kredit semua member dalam tahun periode
    total_transactions = SavingTransactions.objects.filter(
        transaction_date__year=period.year,
        transaction_type__name__icontains='credit',
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

    # Dana yang dialokasikan untuk member (member_services_weight %)
    member_pool = period.total_profit * (period.member_services_weight / Decimal('100'))
    savings_pool = member_pool * (period.total_savings_weight / Decimal('100'))
    transaction_pool = member_pool * (period.total_transaction_weight / Decimal('100'))

    created = 0
    updated = 0
    for member in members:
        member_savings = SavingWallets.objects.filter(
            member=member, deleted_at__isnull=True
        ).aggregate(total=Sum('balance'))['total'] or Decimal('0')

        member_transactions = SavingTransactions.objects.filter(
            member=member,
            transaction_date__year=period.year,
            transaction_type__name__icontains='credit',
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        savings_share = (
            savings_pool * (member_savings / total_savings)
            if total_savings > 0 else Decimal('0')
        )
        transaction_share = (
            transaction_pool * (member_transactions / total_transactions)
            if total_transactions > 0 else Decimal('0')
        )
        total_shu = savings_share + transaction_share

        dist, is_new = ShuMemberDistributions.objects.update_or_create(
            period_year=period.year,
            member=member,
            defaults={
                'total_savings': member_savings,
                'total_shu': total_shu.quantize(Decimal('0.01')),
            },
        )
        if is_new:
            created += 1
        else:
            updated += 1

    period.status = 'calculated'
    period.save(update_fields=['status', 'updated_at'])

    return Response({
        'message': f'SHU tahun {period.year} berhasil dihitung',
        'created': created,
        'updated': updated,
        'total_member_pool': float(member_pool),
    }, status=200)


# ── ADMIN: DISTRIBUTIONS ─────────────────────────────────────────

@api_view(['GET'])
def admin_shu_distributions(request, period_pk):
    """
    Daftar distribusi SHU per member untuk satu periode.
    Query params: ?search=<nama/nik> ?status=pending|approved|paid
    """
    try:
        shu_result = ShuResults.objects.get(pk=period_pk)
        year = shu_result.period_year
    except ShuResults.DoesNotExist:
        return Response([])
    qs = ShuMemberDistributions.objects.filter(period_year=year).select_related('member')

    search = request.query_params.get('search')
    if search:
        qs = qs.filter(
            Q(member__full_name__icontains=search) |
            Q(member__nik_employee__icontains=search)
        )

    status_filter = request.query_params.get('status')
    if status_filter:
        qs = qs.filter(status=status_filter)

    return Response(AdminShuDistributionSerializer(qs.order_by('member__full_name'), many=True).data)


@api_view(['PATCH'])
def admin_shu_distribution_update(request, pk):
    """
    Update status distribusi SHU satu member.
    Body: { status: 'approved'|'paid'|'cancelled', notes? }
    """
    try:
        dist = ShuMemberDistributions.objects.get(pk=pk)
    except ShuMemberDistributions.DoesNotExist:
        return Response({'error': 'Distribusi tidak ditemukan'}, status=404)

    serializer = ShuDistributionUpdateSerializer(data=request.data)
    if not serializer.is_valid():
        return Response(serializer.errors, status=400)

    new_status = serializer.validated_data['status']
    if new_status == 'paid':
        dist.paid_at = timezone.now()
        dist.distributed_status = True
        dist.status_shu = True
    elif new_status == 'approved':
        dist.distributed_status = True
    elif new_status == 'cancelled':
        dist.distributed_status = False
        dist.status_shu = False
    if 'notes' in serializer.validated_data:
        dist.notes = serializer.validated_data['notes']
    dist.updated_at = timezone.now()
    dist.save()

    return Response(AdminShuDistributionSerializer(dist).data)


@api_view(['POST'])
def admin_shu_bulk_approve(request, period_pk):
    """Approve semua distribusi SHU yang masih pending untuk satu periode."""
    try:
        year = ShuResults.objects.get(pk=period_pk).period_year
    except ShuResults.DoesNotExist:
        return Response({'message': '0 distribusi berhasil di-approve'})
    updated = ShuMemberDistributions.objects.filter(
        period_year=year, distributed_status=False
    ).update(distributed_status=True, updated_at=timezone.now())

    return Response({'message': f'{updated} distribusi berhasil di-approve'})


@api_view(['POST'])
def admin_shu_bulk_pay(request, period_pk):
    """Tandai semua distribusi SHU yang sudah approved menjadi paid."""
    now = timezone.now()
    try:
        year = ShuResults.objects.get(pk=period_pk).period_year
    except ShuResults.DoesNotExist:
        return Response({'message': '0 distribusi berhasil dibayarkan'})
    updated = ShuMemberDistributions.objects.filter(
        period_year=year, distributed_status=False
    ).update(paid_at=now, distributed_status=True, status_shu=True, updated_at=now)

    if updated > 0:
        ShuPeriods.objects.filter(pk=period_pk).update(status='distributed', updated_at=now)

    return Response({'message': f'{updated} distribusi berhasil dibayarkan'})


# ── ADMIN: OUTCOME TRANSACTION (income_expenses) ─────────────────

@api_view(['GET'])
def admin_shu_outcome_categories(request):
    """Daftar semua kategori dari income_expense_categories."""
    categories = IncomeExpenseCategories.objects.filter(
        deleted_at__isnull=True
    ).order_by('category_name')
    return Response(IncomeExpenseCategorySerializer(categories, many=True).data)


def _auto_recalculate_shu_results(year: int, month: int):
    """Recalculate and upsert ShuResults for annual (period_month=13) and monthly aggregates."""
    from django.db import transaction as db_txn

    def _upsert(period_year, period_month, filter_kwargs):
        qs = IncomeExpenses.objects.filter(deleted_at__isnull=True, **filter_kwargs)
        agg = qs.aggregate(
            total_income=Sum('amount', filter=Q(category__type__iexact=IncomeExpenses.TYPE_INCOME)),
            total_expense=Sum('amount', filter=Q(category__type__iexact=IncomeExpenses.TYPE_EXPENSE)),
        )
        total_revenue = Decimal(str(agg['total_income'] or 0))
        total_expense_val = Decimal(str(agg['total_expense'] or 0))
        net_profit = (total_revenue - total_expense_val).quantize(Decimal('0.01'))

        with db_txn.atomic():
            result, _ = ShuResults.objects.update_or_create(
                period_year=period_year,
                period_month=period_month,
                defaults={
                    'total_revenue': total_revenue,
                    'total_expense': total_expense_val,
                    'net_profit': net_profit,
                    'distributed_status': True,
                    'deleted_at': None,
                },
            )
            # shu_component_allocations has a DB CHECK period_month BETWEEN 1 AND 12,
            # so skip rebuilding allocations for the annual sentinel (period_month=13).
            if 1 <= period_month <= 12:
                ShuComponentAllocation.objects.filter(shu_result=result).delete()
                active_configs = MasterConfiguration.objects.filter(deleted_at__isnull=True).order_by('id')
                allocations = [
                    ShuComponentAllocation(
                        shu_result=result,
                        master_configuration=config,
                        component_name=config.component_name,
                        percentage=config.percentage,
                        allocated_amount=(result.net_profit * config.percentage / Decimal('100')).quantize(Decimal('0.01')),
                        period_month=result.period_month,
                        period_year=result.period_year,
                    )
                    for config in active_configs
                ]
                if allocations:
                    ShuComponentAllocation.objects.bulk_create(allocations)

    _upsert(year, 13, {'transaction_date__year': year})
    _upsert(year, month, {'transaction_date__year': year, 'transaction_date__month': month})


@api_view(['GET', 'POST'])
def admin_shu_outcome_transactions(request):
    """
    GET  → daftar transaksi outcome dari income_expenses (type='expense').
           Query params: ?search= ?month= ?year=
    POST → tambah transaksi outcome baru ke income_expenses.
    """
    if request.method == 'POST':
        serializer = IncomeExpenseOutcomeCreateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=400)
        txn_obj = serializer.save()
        _auto_recalculate_shu_results(txn_obj.transaction_date.year, txn_obj.transaction_date.month)
        return Response(IncomeExpenseOutcomeSerializer(txn_obj).data, status=201)

    qs = IncomeExpenses.objects.filter(
        deleted_at__isnull=True,
    ).select_related('category').order_by('-transaction_date')

    search = request.query_params.get('search')
    if search:
        qs = qs.filter(
            Q(invoice_number__icontains=search) |
            Q(supplier_customer__icontains=search)
        )

    month = request.query_params.get('month')
    year = request.query_params.get('year')
    day = request.query_params.get('day')
    if month:
        qs = qs.filter(transaction_date__month=month)
    if year:
        qs = qs.filter(transaction_date__year=year)
    if day:
        qs = qs.filter(transaction_date__day=day)

    agg = qs.aggregate(
        total_income=Sum('amount', filter=Q(category__type__iexact=IncomeExpenses.TYPE_INCOME)),
        total_expense=Sum('amount', filter=Q(category__type__iexact=IncomeExpenses.TYPE_EXPENSE)),
    )
    return Response({
        'count': qs.count(),
        'total_income': float(agg['total_income'] or 0),
        'total_expense': float(agg['total_expense'] or 0),
        'results': IncomeExpenseOutcomeSerializer(qs, many=True).data,
    })


@api_view(['GET', 'PATCH', 'DELETE'])
def admin_shu_outcome_transaction_detail(request, pk):
    """GET / PATCH / DELETE satu transaksi dari income_expenses."""
    try:
        txn_obj = IncomeExpenses.objects.select_related('category').get(
            pk=pk, deleted_at__isnull=True
        )
    except IncomeExpenses.DoesNotExist:
        return Response({'error': 'Transaksi tidak ditemukan'}, status=404)

    if request.method == 'GET':
        return Response(IncomeExpenseOutcomeSerializer(txn_obj).data)

    if request.method == 'DELETE':
        old_date = txn_obj.transaction_date
        txn_obj.delete()
        _auto_recalculate_shu_results(old_date.year, old_date.month)
        return Response(status=204)

    old_date = txn_obj.transaction_date
    serializer = IncomeExpenseOutcomeCreateSerializer(txn_obj, data=request.data, partial=True)
    if not serializer.is_valid():
        return Response(serializer.errors, status=400)
    updated = serializer.save()
    new_date = updated.transaction_date
    _auto_recalculate_shu_results(new_date.year, new_date.month)
    if old_date.year != new_date.year or old_date.month != new_date.month:
        _auto_recalculate_shu_results(old_date.year, old_date.month)
    return Response(IncomeExpenseOutcomeSerializer(updated).data)


@api_view(['GET'])
def admin_shu_member_bases(request):
    """
    Daftar anggota beserta simpanan wajib, sukarela, total, dan SHU Jasa Modal
    untuk satu bulan tertentu (default: bulan berjalan) atau ringkasan tahunan.
    Query params: ?search=<nama/nik> ?summary=<month|year> ?month=<1-12> ?year=<yyyy>
    """
    from api.models import Departments
    from api.saving.models import SavingTransactions

    now = timezone.now()
    search = request.query_params.get('search', '')
    summary = request.query_params.get('summary', 'month').lower()
    try:
        year = int(request.query_params.get('year', now.year))
    except (ValueError, TypeError):
        year = now.year

    month_param = request.query_params.get('month')
    try:
        month = int(month_param) if month_param not in (None, '') else None
    except (ValueError, TypeError):
        month = None

    members_qs = Members.objects.filter(deleted_at__isnull=True)
    if search:
        members_qs = members_qs.filter(
            Q(full_name__icontains=search) |
            Q(nik_employee__icontains=search)
        )
    members_qs = members_qs.order_by('full_name')

    departments = {d.id: d.department_name for d in Departments.objects.all()}

    from datetime import date as date_cls
    if summary == 'year':
        period_start = date_cls(year, 1, 1)
        period_end = date_cls(year, 12, 31)
        # Jumlah bulan yang dihitung: Januari s/d bulan sekarang (jika tahun berjalan)
        # atau 12 (jika tahun sudah lewat).
        if year < now.year:
            months_multiplier = Decimal('12')
        elif year == now.year:
            months_multiplier = Decimal(str(now.month))
        else:
            months_multiplier = Decimal('0')
    else:
        period_month = month or now.month
        period_start = date_cls(year, period_month, 1)
        period_end = date_cls(year, period_month, 28)
        months_multiplier = Decimal('1')

    # If obligation values are defined on the member_saving_obligations table,
    # use those monthly amounts directly for mandatory/voluntary savings.
    obligations = MemberSavingObligations.objects.filter(
        member_id__in=members_qs.values_list('id', flat=True),
        saving_type_id__in=[1, 2],
        is_active=True,
        deleted_at__isnull=True,
    ).values('member_id', 'saving_type_id', 'monthly_amount')

    member_savings = {}
    for o in obligations:
        mid = o['member_id']
        if mid not in member_savings:
            member_savings[mid] = {'mandatory': Decimal('0'), 'voluntary': Decimal('0')}
        if o['saving_type_id'] == 1:
            member_savings[mid]['mandatory'] += o['monthly_amount'] * months_multiplier
        elif o['saving_type_id'] == 2:
            member_savings[mid]['voluntary'] += o['monthly_amount'] * months_multiplier

    # Total simpanan semua member aktif (denominator untuk proporsi jasa modal)
    all_obligations_agg = MemberSavingObligations.objects.filter(
        saving_type_id__in=[1, 2],
        is_active=True,
        deleted_at__isnull=True,
        member__deleted_at__isnull=True,
    ).aggregate(total=Sum('monthly_amount'))
    total_all_savings = (all_obligations_agg['total'] or Decimal('0')) * months_multiplier

    # Hitung jasa modal pool dari shu_component_allocations (master_configuration_id=1)
    jasa_modal_pool = None
    try:
        period_month_for_result = (month or now.month) if summary == 'month' else 13
        shu_result = ShuResults.objects.get(
            period_year=year,
            period_month=period_month_for_result,
            deleted_at__isnull=True,
        )
        try:
            jasa_modal_alloc = ShuComponentAllocation.objects.get(
                shu_result=shu_result,
                master_configuration_id=1,
                deleted_at__isnull=True,
            )
            jasa_modal_pool = jasa_modal_alloc.allocated_amount
        except ShuComponentAllocation.DoesNotExist:
            jasa_modal_cfg = MasterConfiguration.objects.filter(
                pk=1, deleted_at__isnull=True,
            ).first()
            if jasa_modal_cfg:
                jasa_modal_pool = (
                    shu_result.net_profit * jasa_modal_cfg.percentage / Decimal('100')
                ).quantize(Decimal('0.01'))
    except ShuResults.DoesNotExist:
        pass

    results = []
    for m in members_qs:
        savings = member_savings.get(m.id, {'mandatory': Decimal('0'), 'voluntary': Decimal('0')})
        total = savings['mandatory'] + savings['voluntary']

        shu_jasa_modal = None
        if jasa_modal_pool is not None and total_all_savings > 0:
            shu_jasa_modal = float(
                (total / total_all_savings * jasa_modal_pool).quantize(Decimal('0.01'))
            )

        results.append({
            'member_id': m.id,
            'full_name': m.full_name,
            'nik_employee': m.nik_employee or '-',
            'department_name': departments.get(m.department_id, '-'),
            'mandatory_saving_monthly': float(savings['mandatory']),
            'voluntary_saving_monthly': float(savings['voluntary']),
            'total_saving_amount': float(total),
            'shu_jasa_modal': shu_jasa_modal,
        })

    return Response({
        'count': len(results),
        'month': month,
        'year': year,
        'summary': summary,
        'months_count': int(months_multiplier),
        'jasa_modal_pool': float(jasa_modal_pool) if jasa_modal_pool is not None else None,
        'total_all_savings': float(total_all_savings),
        'results': results,
    })


@api_view(['POST'])
def admin_shu_member_bases_distribute(request):
    """
    Persist current calculated `shu_jasa_modal` and `total_saving_amount` per member
    into the `shu_member_bases` table linked to a `shu_result` (by year/month).
    Body: { year: <yyyy>, month?: <1-12> }
    """
    from api.models import Departments

    now = timezone.now()
    try:
        year = int(request.data.get('year', now.year))
    except (TypeError, ValueError):
        return Response({'error': 'year tidak valid'}, status=400)

    month_param = request.data.get('month')
    try:
        month = int(month_param) if month_param not in (None, '') else None
    except (ValueError, TypeError):
        return Response({'error': 'month tidak valid'}, status=400)

    # find shu_result
    try:
        shu_result = ShuResults.objects.get(period_year=year, period_month=month, deleted_at__isnull=True)
    except ShuResults.DoesNotExist:
        return Response({'error': 'ShuResults untuk periode tidak ditemukan'}, status=404)

    # Build obligations and totals (same logic as admin_shu_member_bases)
    obligations = MemberSavingObligations.objects.filter(
        saving_type_id__in=[1, 2],
        is_active=True,
        deleted_at__isnull=True,
        member__deleted_at__isnull=True,
    ).values('member_id', 'saving_type_id', 'monthly_amount')

    member_savings = {}
    for o in obligations:
        mid = o['member_id']
        if mid not in member_savings:
            member_savings[mid] = {'mandatory': Decimal('0'), 'voluntary': Decimal('0')}
        if o['saving_type_id'] == 1:
            member_savings[mid]['mandatory'] += o['monthly_amount']
        elif o['saving_type_id'] == 2:
            member_savings[mid]['voluntary'] += o['monthly_amount']

    total_all_savings = MemberSavingObligations.objects.filter(
        saving_type_id__in=[1, 2],
        is_active=True,
        deleted_at__isnull=True,
        member__deleted_at__isnull=True,
    ).aggregate(total=Sum('monthly_amount'))['total'] or Decimal('0')

    jasa_modal_cfg = MasterConfiguration.objects.filter(
        component_name__icontains='jasa modal',
        deleted_at__isnull=True,
    ).first()

    if not jasa_modal_cfg:
        return Response({'error': 'Konfigurasi komponen "Jasa Modal" belum diset'}, status=400)

    jasa_modal_pool = (shu_result.net_profit * jasa_modal_cfg.percentage / Decimal('100')).quantize(Decimal('0.01'))

    members = Members.objects.filter(deleted_at__isnull=True)

    created = 0
    updated = 0
    with transaction.atomic():
        for m in members:
            savings = member_savings.get(m.id, {'mandatory': Decimal('0'), 'voluntary': Decimal('0')})
            total = savings['mandatory'] + savings['voluntary']

            shu_jasa_modal = Decimal('0')
            if total_all_savings > 0:
                shu_jasa_modal = (total / total_all_savings * jasa_modal_pool).quantize(Decimal('0.01'))

            try:
                obj, is_new = ShuMemberBases.objects.update_or_create(
                    shu_result=shu_result,
                    member=m,
                    defaults={
                        'total_saving_amount': total,
                        'shu_jasa_modal': shu_jasa_modal,
                    },
                )
                if is_new:
                    created += 1
                else:
                    updated += 1
            except IntegrityError as exc:
                # Try to recover from Postgres sequence mismatch (duplicate PK)
                msg = str(exc)
                if 'duplicate key value violates unique constraint' in msg and 'pkey' in msg:
                    try:
                        seq_sql = "SELECT pg_get_serial_sequence('shu_member_bases','id')"
                        with connection.cursor() as cur:
                            cur.execute(seq_sql)
                            seq = cur.fetchone()[0]
                            if seq:
                                # set sequence to max(id)
                                cur.execute(
                                    "SELECT setval(%s, COALESCE((SELECT MAX(id) FROM shu_member_bases), 0))",
                                    [seq],
                                )
                        # retry once
                        obj, is_new = ShuMemberBases.objects.update_or_create(
                            shu_result=shu_result,
                            member=m,
                            defaults={
                                'total_saving_amount': total,
                                'shu_jasa_modal': shu_jasa_modal,
                            },
                        )
                        if is_new:
                            created += 1
                        else:
                            updated += 1
                        continue
                    except Exception as exc2:
                        return Response({'error': 'IntegrityError', 'details': msg, 'recovery_error': str(exc2)}, status=500)
                return Response({'error': 'IntegrityError', 'details': msg}, status=500)

    return Response({'message': 'Data shu_member_bases disimpan', 'created': created, 'updated': updated})


@api_view(['GET', 'POST'])
def admin_shu_results(request):
    """
    GET  → cek hasil SHU tersimpan untuk periode tertentu.
           Query params: ?year=<yyyy> &month=<1-12>
    POST → simpan/update hasil SHU ke tabel shu_results.
           Body: { period_year, period_month?, total_revenue, total_expense, net_profit }
    """
    if request.method == 'GET':
        qs = ShuResults.objects.filter(deleted_at__isnull=True)
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        if year:
            qs = qs.filter(period_year=year)
        if month:
            qs = qs.filter(period_month=month)
        else:
            qs = qs.filter(period_month=13)
        result = qs.first()
        if result:
            return Response(ShuResultsSerializer(result).data)
        return Response({})

    # POST
    year = request.data.get('period_year')
    month = request.data.get('period_month') or None

    if not year:
        return Response({'error': 'period_year wajib diisi'}, status=400)

    # Try to read provided totals; if not provided (or zero), compute from IncomeExpenses
    provided_revenue = request.data.get('total_revenue', None)
    provided_expense = request.data.get('total_expense', None)
    provided_net = request.data.get('net_profit', None)

    try:
        if provided_revenue is not None:
            total_revenue = Decimal(str(provided_revenue))
        else:
            total_revenue = None
        if provided_expense is not None:
            total_expense = Decimal(str(provided_expense))
        else:
            total_expense = None
        if provided_net is not None:
            net_profit = Decimal(str(provided_net))
        else:
            net_profit = None
    except (InvalidOperation, TypeError):
        return Response({'error': 'Nilai total_revenue/total_expense/net_profit tidak valid'}, status=400)

    if total_revenue is None or total_expense is None or net_profit is None:
        # Compute aggregates from IncomeExpenses for the given period
        qs = IncomeExpenses.objects.filter(deleted_at__isnull=True)
        if month:
            qs = qs.filter(transaction_date__year=int(year), transaction_date__month=int(month))
        else:
            qs = qs.filter(transaction_date__year=int(year))

        agg = qs.aggregate(
            total_income=Sum('amount', filter=Q(category__type__iexact=IncomeExpenses.TYPE_INCOME)),
            total_expense=Sum('amount', filter=Q(category__type__iexact=IncomeExpenses.TYPE_EXPENSE)),
        )

        computed_income = agg['total_income'] or 0
        computed_expense = agg['total_expense'] or 0

        if total_revenue is None:
            total_revenue = Decimal(str(computed_income))
        if total_expense is None:
            total_expense = Decimal(str(computed_expense))
        if net_profit is None:
            net_profit = (total_revenue - total_expense).quantize(Decimal('0.01'))

    with transaction.atomic():
        result, created = ShuResults.objects.update_or_create(
            period_year=int(year),
            period_month=int(month) if month else 13,
            defaults={
                'total_revenue': total_revenue,
                'total_expense': total_expense,
                'net_profit': net_profit,
                'distributed_status': True,
            },
        )

        # Clear existing allocations and rebuild them from MasterConfiguration
        ShuComponentAllocation.objects.filter(shu_result=result).delete()
        active_configs = MasterConfiguration.objects.filter(deleted_at__isnull=True).order_by('id')
        allocations = []
        for config in active_configs:
            allocated_amount = (result.net_profit * config.percentage / Decimal('100')).quantize(Decimal('0.01'))
            allocations.append(ShuComponentAllocation(
                shu_result=result,
                master_configuration=config,
                component_name=config.component_name,
                percentage=config.percentage,
                allocated_amount=allocated_amount,
                period_month=result.period_month,
                period_year=result.period_year,
            ))
        if allocations:
            ShuComponentAllocation.objects.bulk_create(allocations)

    return Response(ShuResultsSerializer(result).data, status=201 if created else 200)


def _build_month_labels(year, month, count):
    labels = []
    values = []
    total_months = year * 12 + month - 1
    for offset in range(count - 1, -1, -1):
        index = total_months - offset
        y = index // 12
        m = index % 12 + 1
        labels.append(f'{y}-{m:02d}')
        values.append((y, m))
    return labels, values


@api_view(['GET'])
def admin_shu_net_sales(request):
    """Return SHU (net_profit) series from shu_results grouped by period_year/period_month."""
    from datetime import date

    range_option = request.query_params.get('range', '3month')
    range_map = {
        '1month': 1,
        '3month': 3,
        '6month': 6,
        '1year': 12,
        '3year': 36,
    }
    months = range_map.get(range_option, 6)
    today = timezone.now().date()
    labels, month_pairs = _build_month_labels(today.year, today.month, months)
    start_year, start_month = month_pairs[0]
    start_code = start_year * 100 + start_month

    query = """
        SELECT period_year, period_month, COALESCE(net_profit, 0)
        FROM shu_results
        WHERE deleted_at IS NULL
          AND period_month BETWEEN 1 AND 12
          AND (period_year * 100 + period_month) >= %s
        ORDER BY period_year, period_month
    """

    with connection.cursor() as cursor:
        cursor.execute(query, [start_code])
        rows = cursor.fetchall()

    result_map = {
        f'{row[0]}-{row[1]:02d}': float(row[2])
        for row in rows
    }

    data = [result_map.get(key, 0) for key in labels]
    readable_labels = []
    for key in labels:
        year_str, month_str = key.split('-')
        readable_labels.append(f'{date(int(year_str), int(month_str), 1):%b %Y}')

    return Response({
        'range': range_option,
        'labels': readable_labels,
        'data': data,
    })


@api_view(['GET'])
def admin_shu_weekly_cashflow(request):
    """
    Return weekly money-in / money-out cashflow using the same data sources
    as the transaction_history endpoint:
      Money IN  → saving_transactions (deposits) + loan_payments (installments)
      Money OUT → withdrawals + shu_member_distributions (SHU distributions)
    """
    from datetime import timedelta

    range_option = request.query_params.get('range', '3month')
    range_map = {
        '7days': 7,
        '30days': 30,
        'weekly': 28,
        '3month': 90,
        '6month': 180,
        '1year': 365,
        '3year': 1095,
    }
    days = range_map.get(range_option, 90)
    today = timezone.now().date()
    start_date = today - timedelta(days=days)

    if range_option in ['7days', '30days']:
        trunc_unit = 'day'
        step_days = 1
        start_period = start_date
    else:
        trunc_unit = 'week'
        step_days = 7
        start_period = start_date - timedelta(days=start_date.weekday())

    query = f"""
        SELECT
            DATE_TRUNC('{trunc_unit}', transaction_date)::date AS period_start,
            SUM(CASE WHEN flow = 'IN'  THEN amount ELSE 0 END) AS total_income,
            SUM(CASE WHEN flow = 'OUT' THEN amount ELSE 0 END) AS total_expense
        FROM (
            -- MONEY IN: saving deposits (MANDATORY / VOLUNTARY / PRINCIPLE / DEPOSIT)
            SELECT
                st.transaction_date::date AS transaction_date,
                st.amount,
                'IN' AS flow
            FROM saving_transactions st
            INNER JOIN transaction_types tt ON tt.id = st.transaction_type_id
            INNER JOIN statuses s ON s.id = st.status_id
            WHERE st.transaction_date::date >= %s
              AND st.transaction_date::date <= %s
              AND UPPER(tt.name) IN ('MANDATORY', 'VOLUNTARY', 'PRINCIPLE', 'DEPOSIT', 'CREDIT')
              AND UPPER(s.status_code) IN ('COMPLETED', 'PAID', 'SUCCESS')

            UNION ALL

            -- MONEY IN: loan installment payments
            SELECT
                lp.payment_date::date AS transaction_date,
                lp.amount_paid AS amount,
                'IN' AS flow
            FROM loan_payments lp
            INNER JOIN statuses s ON s.id = lp.status_id
            WHERE lp.payment_date::date >= %s
              AND lp.payment_date::date <= %s
              AND UPPER(s.status_code) IN ('COMPLETED', 'PAID', 'LATE_PAID', 'SUCCESS')

            UNION ALL

            -- MONEY OUT: withdrawals
            SELECT
                w.request_date::date AS transaction_date,
                w.amount,
                'OUT' AS flow
            FROM withdrawals w
            INNER JOIN statuses s ON s.id = w.status_id
            WHERE w.request_date::date >= %s
              AND w.request_date::date <= %s
              AND UPPER(s.status_code) IN ('COMPLETED', 'PAID', 'SUCCESS')

            UNION ALL

            -- MONEY OUT: SHU distributions
             SELECT
                smd.paid_at::date AS transaction_date,
                smd.total_shu AS amount,
                'OUT' AS flow
            FROM shu_member_distributions smd
            WHERE smd.paid_at::date >= %s
                AND smd.paid_at::date <= %s
                AND smd.distributed_status = TRUE
                AND smd.status_shu = TRUE

        ) combined
        WHERE transaction_date IS NOT NULL
        GROUP BY DATE_TRUNC('{trunc_unit}', transaction_date)::date
        ORDER BY period_start
    """

    params = [
        start_period, today,  # saving deposits
        start_period, today,  # installments
        start_period, today,  # withdrawals
        start_period, today,  # SHU distributions
    ]

    period_map = {}
    try:
        with connection.cursor() as cursor:
            cursor.execute(query, params)
            for row in cursor.fetchall():
                period_start_date, total_in, total_out = row
                if period_start_date:
                    period_map[period_start_date] = {
                        'income': float(total_in or 0),
                        'expense': float(total_out or 0),
                    }
    except Exception as e:
        print("ERROR IN admin_shu_weekly_cashflow:", e)
        pass

    labels = []
    income_data = []
    expense_data = []
    cursor_date = start_period
    while cursor_date <= today:
        labels.append(cursor_date.strftime('%d %b'))
        totals = period_map.get(cursor_date, {'income': 0, 'expense': 0})
        income_data.append(totals['income'])
        expense_data.append(totals['expense'])
        cursor_date += timedelta(days=step_days)

    return Response({
        'range': range_option,
        'labels': labels,
        'income': income_data,
        'expense': expense_data,
    })


@api_view(['GET'])
def admin_shu_outcome_excel_template(request):
    """Generate dan return Excel template untuk upload transaksi ke income_expenses."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return Response({'error': 'openpyxl belum terinstall. Jalankan: pip install openpyxl'}, status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Template'

    header_labels = [
        'transaction_date (YYYY-MM-DD)',
        'category_id',
        'invoice_number',
        'supplier_customer',
        'quantity',
        'amount',
    ]
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    for col, label in enumerate(header_labels, 1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Contoh baris
    sample_row = ['2025-01-15', 1, 'INV-31012025001', 'PT. Supplier ABC', 10, 500000]
    for col, val in enumerate(sample_row, 1):
        ws.cell(row=2, column=col, value=val)

    for col, width in enumerate([28, 12, 22, 30, 12, 18], 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # Sheet daftar kategori sebagai referensi
    ws2 = wb.create_sheet('Daftar Kategori')
    for col, label in enumerate(['ID', 'Nama Kategori', 'Type (INCOME/EXPENSE)'], 1):
        cell = ws2.cell(row=1, column=col, value=label)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid')

    categories = IncomeExpenseCategories.objects.filter(deleted_at__isnull=True).order_by('type', 'category_name')
    for row_num, cat in enumerate(categories, 2):
        ws2.cell(row=row_num, column=1, value=cat.id)
        ws2.cell(row=row_num, column=2, value=cat.category_name)
        ws2.cell(row=row_num, column=3, value=cat.type)

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 35
    ws2.column_dimensions['C'].width = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="template_transaksi_shu.xlsx"'
    wb.save(response)
    return response


@api_view(['POST'])
def admin_shu_outcome_upload_excel(request):
    """
    Upload file Excel dan bulk-insert ke tabel income_expenses.
    Field per baris: transaction_date, category_id, invoice_number, supplier_customer, quantity, amount.
    """
    try:
        import openpyxl
    except ImportError:
        return Response({'error': 'openpyxl belum terinstall. Jalankan: pip install openpyxl'}, status=500)

    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'File tidak ditemukan dalam request.'}, status=400)

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as exc:
        return Response({'error': f'File Excel tidak dapat dibaca: {exc}'}, status=400)

    errors = []
    records = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell is not None for cell in row):
            continue

        cells = list(row) + [None] * 6
        raw_date, raw_cat_id, raw_invoice, raw_supcust, raw_qty, raw_amount = cells[:6]

        # Validasi kolom wajib
        if raw_date is None or raw_cat_id is None or raw_amount is None:
            errors.append(f'Baris {row_num}: kolom transaction_date, category_id, dan amount wajib diisi.')
            continue

        # Parse tanggal — bisa datetime (dari Excel) atau string
        try:
            if isinstance(raw_date, datetime):
                parsed_date = raw_date.date()
            else:
                parsed_date = datetime.strptime(str(raw_date).strip(), '%Y-%m-%d').date()
        except ValueError:
            errors.append(f'Baris {row_num}: format tanggal tidak valid "{raw_date}" (harus YYYY-MM-DD).')
            continue

        # Validasi category
        try:
            category = IncomeExpenseCategories.objects.get(pk=int(raw_cat_id), deleted_at__isnull=True)
        except (IncomeExpenseCategories.DoesNotExist, ValueError, TypeError):
            errors.append(f'Baris {row_num}: category_id "{raw_cat_id}" tidak ditemukan.')
            continue

        # Parse amount & quantity
        try:
            amount = Decimal(str(raw_amount))
        except InvalidOperation:
            errors.append(f'Baris {row_num}: nilai amount "{raw_amount}" tidak valid.')
            continue

        try:
            quantity = Decimal(str(raw_qty)) if raw_qty is not None else None
        except InvalidOperation:
            quantity = None

        records.append(IncomeExpenses(
            transaction_date=parsed_date,
            category=category,
            type=category.type,
            invoice_number=str(raw_invoice).strip() if raw_invoice is not None else None,
            supplier_customer=str(raw_supcust).strip() if raw_supcust is not None else None,
            quantity=quantity,
            amount=amount,
            created_at=timezone.now(),
            updated_at=timezone.now(),
        ))

    if not records and errors:
        return Response({'inserted': 0, 'errors': errors}, status=400)

    IncomeExpenses.objects.bulk_create(records)

    # Auto-recalculate ShuResults for every affected year+month
    affected_periods = set()
    for r in records:
        d = r.transaction_date
        affected_periods.add((d.year, d.month))
    for yr, mo in affected_periods:
        _auto_recalculate_shu_results(yr, mo)

    return Response({
        'inserted': len(records),
        'errors': errors,
    }, status=201)


# ── ADMIN: ANNUAL JASA MODAL DISTRIBUTION ────────────────────────

@api_view(['GET'])
def admin_shu_jasa_modal_list(request):
    """
    Daftar distribusi SHU Jasa Modal tahunan per member beserta info bank.
    Query params: ?year=<yyyy> &search=<nama/nik>
    """
    year_param = request.query_params.get('year')
    if not year_param:
        return Response({'error': 'year wajib diisi'}, status=400)

    try:
        year_int = int(year_param)
    except ValueError:
        return Response({'error': 'year tidak valid'}, status=400)

    try:
        period = ShuResults.objects.get(period_year=year_int, period_month=13, deleted_at__isnull=True)
    except ShuResults.DoesNotExist:
        return Response({'distributed': False, 'results': []})

    qs = ShuMemberDistributions.objects.filter(
        period_year=year_int
    ).select_related('member').order_by('member__full_name')

    search = request.query_params.get('search', '')
    if search:
        qs = qs.filter(
            Q(member__full_name__icontains=search) |
            Q(member__nik_employee__icontains=search)
        )

    member_ids = list(qs.values_list('member_id', flat=True))
    bank_map = _get_bank_map(member_ids)

    from api.models import Departments
    departments = {d.id: d.department_name for d in Departments.objects.all()}

    return Response({
        'distributed': True,
        'period_id': period.id,
        'results': AdminAnnualJasaModalSerializer(
            qs, many=True, context={'bank_map': bank_map, 'departments': departments}
        ).data,
    })


@api_view(['GET'])
def admin_shu_annual_from_monthly(request):
    """
    Agregasi data dari shu_member_distributions_monthly per tahun.
    Menjumlahkan simp_wajib, simp_sukarela, total_savings, total_shu untuk semua bulan di tahun yang dipilih.
    Query params: ?year=<yyyy> &search=<nama/nik>
    """
    from api.models import Departments

    now = timezone.now()
    try:
        year = int(request.query_params.get('year', now.year))
    except (ValueError, TypeError):
        return Response({'error': 'year tidak valid'}, status=400)

    monthly_agg = (
        ShuMemberDistributionsMonthly.objects
        .filter(period_year=year)
        .values('member_id')
        .annotate(
            total_simp_wajib=Sum('simp_wajib'),
            total_simp_sukarela=Sum('simp_sukarela'),
            total_simpanan=Sum('total_savings'),
            total_shu_sum=Sum('total_shu'),
        )
    )

    member_data = {
        row['member_id']: {
            'simp_wajib': row['total_simp_wajib'] or Decimal('0'),
            'simp_sukarela': row['total_simp_sukarela'] or Decimal('0'),
            'total_savings': row['total_simpanan'] or Decimal('0'),
            'total_shu': row['total_shu_sum'] or Decimal('0'),
        }
        for row in monthly_agg
    }

    search = request.query_params.get('search', '')
    members_qs = Members.objects.filter(deleted_at__isnull=True)
    if search:
        members_qs = members_qs.filter(
            Q(full_name__icontains=search) |
            Q(nik_employee__icontains=search)
        )
    members_qs = members_qs.order_by('full_name')

    departments = {d.id: d.department_name for d in Departments.objects.all()}

    all_member_ids = list(members_qs.values_list('id', flat=True))
    bank_map = _get_bank_map(all_member_ids)

    results = []
    for m in members_qs:
        d = member_data.get(m.id, {
            'simp_wajib': Decimal('0'),
            'simp_sukarela': Decimal('0'),
            'total_savings': Decimal('0'),
            'total_shu': Decimal('0'),
        })
        bank_info = bank_map.get(m.id)

        results.append({
            'member_id': m.id,
            'full_name': m.full_name,
            'nik_employee': m.nik_employee or '-',
            'department_name': departments.get(m.department_id, '-'),
            'simp_wajib': float(d['simp_wajib']),
            'simp_sukarela': float(d['simp_sukarela']),
            'total_savings': float(d['total_savings']),
            'total_shu': float(d['total_shu']),
            'bank_info': bank_info,
        })

    has_data = any(r['total_savings'] > 0 or r['total_shu'] > 0 for r in results)
    total_shu_pool = sum(r['total_shu'] for r in results)

    return Response({
        'count': len(results),
        'year': year,
        'has_data': has_data,
        'total_shu_pool': total_shu_pool if has_data else None,
        'results': results,
    })


@api_view(['POST'])
def admin_shu_jasa_modal_distribute(request):
    """
    Distribusikan SHU Jasa Modal tahunan ke seluruh member.
    Mengambil data dari shu_member_distributions_monthly (sum per tahun) lalu disimpan ke shu_member_distributions.
    Body: { year: <yyyy> }
    """
    now = timezone.now()
    try:
        year = int(request.data.get('year', 0))
    except (TypeError, ValueError):
        return Response({'error': 'year tidak valid'}, status=400)

    if not year:
        return Response({'error': 'year wajib diisi'}, status=400)

    if year > now.year:
        return Response({'error': 'Tidak dapat mendistribusikan SHU untuk tahun yang akan datang'}, status=400)

    try:
        shu_result = ShuResults.objects.get(
            period_year=year, period_month=13, deleted_at__isnull=True
        )
    except ShuResults.DoesNotExist:
        return Response(
            {'error': f'SHU Result tahunan untuk tahun {year} belum ada. Buat SHU Result terlebih dahulu.'},
            status=404,
        )

    # Aggregate from monthly distributions table
    monthly_agg = (
        ShuMemberDistributionsMonthly.objects
        .filter(period_year=year)
        .values('member_id')
        .annotate(
            total_simp_wajib=Sum('simp_wajib'),
            total_simp_sukarela=Sum('simp_sukarela'),
            total_simpanan=Sum('total_savings'),
            total_shu_sum=Sum('total_shu'),
        )
    )

    member_data = {
        row['member_id']: {
            'simp_wajib': row['total_simp_wajib'] or Decimal('0'),
            'simp_sukarela': row['total_simp_sukarela'] or Decimal('0'),
            'total_savings': row['total_simpanan'] or Decimal('0'),
            'total_shu': row['total_shu_sum'] or Decimal('0'),
        }
        for row in monthly_agg
    }

    if not member_data:
        return Response(
            {'error': f'Belum ada data distribusi bulanan untuk tahun {year}. Distribusikan SHU bulanan terlebih dahulu.'},
            status=404,
        )

    # Optional: only distribute to a specific subset of member IDs
    member_ids = request.data.get('member_ids')
    members = Members.objects.filter(deleted_at__isnull=True)
    if member_ids:
        members = members.filter(id__in=member_ids)
    created = updated = 0
    now_ts = timezone.now()

    try:
        with transaction.atomic():
            for m in members:
                d = member_data.get(m.id, {
                    'simp_wajib': Decimal('0'),
                    'simp_sukarela': Decimal('0'),
                    'total_savings': Decimal('0'),
                    'total_shu': Decimal('0'),
                })
                try:
                    dist = ShuMemberDistributions.objects.get(period_year=year, member=m)
                    dist.simp_wajib = d['simp_wajib']
                    dist.simp_sukarela = d['simp_sukarela']
                    dist.total_savings = d['total_savings']
                    dist.total_shu = d['total_shu']
                    dist.status_shu = True
                    dist.updated_at = now_ts
                    dist.save(update_fields=['simp_wajib', 'simp_sukarela', 'total_savings', 'total_shu', 'status_shu', 'updated_at'])
                    updated += 1
                except ShuMemberDistributions.DoesNotExist:
                    ShuMemberDistributions.objects.create(
                        member=m,
                        simp_wajib=d['simp_wajib'],
                        simp_sukarela=d['simp_sukarela'],
                        total_savings=d['total_savings'],
                        total_shu=d['total_shu'],
                        status_shu=True,
                        period_year=year,
                        created_at=now_ts,
                        updated_at=now_ts,
                    )
                    created += 1
    except Exception as exc:
        return Response({'error': f'[shu_member_distributions] {type(exc).__name__}: {exc}'}, status=500)

    return Response({
        'message': f'SHU Jasa Modal tahun {year} berhasil didistribusikan',
        'created': created,
        'updated': updated,
    })


@api_view(['PATCH'])
@parser_classes([MultiPartParser, FormParser])
def admin_shu_jasa_modal_proof_upload(request, pk):
    """
    Upload bukti transfer untuk satu distribusi SHU Jasa Modal.
    Menyimpan URL ke kolom transfer_proof dan mengubah status menjadi PAID (id=39).
    """
    import os
    import uuid
    import boto3

    try:
        dist = ShuMemberDistributions.objects.select_related('member').get(pk=pk)
    except ShuMemberDistributions.DoesNotExist:
        return Response({'error': 'Distribusi tidak ditemukan'}, status=404)

    file = request.FILES.get('file')
    if not file:
        return Response({'error': 'File tidak ditemukan dalam request'}, status=400)

    SUPABASE_URL = os.getenv('SUPABASE_URL', '').rstrip('/')
    S3_BUCKET = os.getenv('AWS_STORAGE_BUCKET_NAME', 'koperasi')
    S3_FOLDER = 'shu/proofs'

    ext = os.path.splitext(file.name)[1].lower()
    s3_key = f"{S3_FOLDER}/{uuid.uuid4().hex}{ext}"

    try:
        s3 = boto3.client(
            's3',
            endpoint_url=os.getenv('AWS_S3_ENDPOINT_URL'),
            aws_access_key_id=os.getenv('AWS_ACCESS_KEY_ID'),
            aws_secret_access_key=os.getenv('AWS_SECRET_ACCESS_KEY'),
            region_name=os.getenv('AWS_S3_REGION_NAME', 'ap-northeast-1'),
        )
        s3.upload_fileobj(
            file,
            S3_BUCKET,
            s3_key,
            ExtraArgs={'ContentType': file.content_type or 'application/octet-stream'},
        )
    except Exception as exc:
        return Response({'error': f'Gagal upload file: {exc}'}, status=500)

    public_url = f"{SUPABASE_URL}/storage/v1/object/public/{S3_BUCKET}/{s3_key}"

    with connection.cursor() as cur:
        cur.execute("SELECT sp_generate_shu_transfer_reference()")
        tf_ref = cur.fetchone()[0]

    dist.transfer_proof = public_url
    dist.transfer_proof_url = public_url
    dist.transfer_proof_name = file.name
    dist.tf_reference_id = tf_ref
    dist.paid_at = timezone.now()
    dist.distributed_status = True
    dist.status_shu = True
    dist.save(update_fields=[
        'transfer_proof', 'transfer_proof_url', 'transfer_proof_name',
        'tf_reference_id', 'paid_at', 'distributed_status', 'status_shu', 'updated_at',
    ])

    from api.saving.email_utils import send_shu_paid_email
    send_shu_paid_email(dist, public_url)

    bank_map = _get_bank_map([dist.member_id])

    from api.models import Departments
    departments = {d.id: d.department_name for d in Departments.objects.all()}

    return Response(AdminAnnualJasaModalSerializer(dist, context={'bank_map': bank_map, 'departments': departments}).data)


@api_view(['PATCH'])
def admin_shu_jasa_modal_update_notes(request, pk):
    """Update notes for one annual jasa modal distribution."""
    try:
        dist = ShuMemberDistributions.objects.select_related('member').get(pk=pk)
    except ShuMemberDistributions.DoesNotExist:
        return Response({'error': 'Distribusi tidak ditemukan'}, status=404)

    dist.notes = request.data.get('notes', '') or None
    dist.updated_at = timezone.now()
    dist.save(update_fields=['notes', 'updated_at'])

    bank_map = _get_bank_map([dist.member_id])

    from api.models import Departments
    departments = {d.id: d.department_name for d in Departments.objects.all()}

    return Response(AdminAnnualJasaModalSerializer(dist, context={'bank_map': bank_map, 'departments': departments}).data)


def _sync_annual_distributions(year: int):
    """
    Agregasi data dari shu_member_distributions_monthly untuk semua bulan di tahun tertentu,
    lalu upsert ke shu_member_distributions (tahunan).
    Hanya berjalan jika ShuResults period_month=13 untuk tahun tersebut sudah ada.
    """
    try:
        annual_result = ShuResults.objects.get(period_year=year, period_month=13, deleted_at__isnull=True)
    except ShuResults.DoesNotExist:
        return  # Annual SHU result belum dibuat, skip

    monthly_agg = (
        ShuMemberDistributionsMonthly.objects
        .filter(period_year=year)
        .values('member_id')
        .annotate(
            total_simp_wajib=Sum('simp_wajib'),
            total_simp_sukarela=Sum('simp_sukarela'),
            total_simpanan=Sum('total_savings'),
            total_shu_sum=Sum('total_shu'),
        )
    )

    if not monthly_agg:
        return

    now_ts = timezone.now()
    for row in monthly_agg:
        simp_wajib = row['total_simp_wajib'] or Decimal('0')
        simp_sukarela = row['total_simp_sukarela'] or Decimal('0')
        total_savings = row['total_simpanan'] or Decimal('0')
        total_shu = row['total_shu_sum'] or Decimal('0')
        try:
            dist = ShuMemberDistributions.objects.get(period_year=year, member_id=row['member_id'])
            dist.simp_wajib = simp_wajib
            dist.simp_sukarela = simp_sukarela
            dist.total_savings = total_savings
            dist.total_shu = total_shu
            dist.updated_at = now_ts
            dist.save(update_fields=['simp_wajib', 'simp_sukarela', 'total_savings', 'total_shu', 'updated_at'])
        except ShuMemberDistributions.DoesNotExist:
            ShuMemberDistributions.objects.create(
                member_id=row['member_id'],
                simp_wajib=simp_wajib,
                simp_sukarela=simp_sukarela,
                total_savings=total_savings,
                total_shu=total_shu,
                period_year=year,
                created_at=now_ts,
                updated_at=now_ts,
            )


@api_view(['POST'])
def admin_shu_monthly_distribute(request):
    """
    Distribusikan SHU Jasa Modal bulanan ke seluruh anggota aktif.
    Hasil disimpan ke tabel shu_member_distributions_monthly.
    Body: { year: <yyyy>, month: <1-12> }

    Mapping:
      period_id  → shu_results.id  (period_year=year, period_month=month)
      member_id  → members.id
      total_savings → total simpanan (wajib + sukarela) per anggota bulan itu
      total_shu     → SHU jasa modal per anggota
    """
    now = timezone.now()
    try:
        year = int(request.data.get('year', now.year))
    except (TypeError, ValueError):
        return Response({'error': 'year tidak valid'}, status=400)

    try:
        month_raw = request.data.get('month')
        month = int(month_raw) if month_raw not in (None, '') else now.month
    except (TypeError, ValueError):
        return Response({'error': 'month tidak valid'}, status=400)

    if not (1 <= month <= 12):
        return Response({'error': 'month harus antara 1 dan 12'}, status=400)

    # Pastikan ShuResults untuk bulan ini sudah ada
    try:
        shu_result = ShuResults.objects.get(
            period_year=year,
            period_month=month,
            deleted_at__isnull=True,
        )
    except ShuResults.DoesNotExist:
        return Response(
            {'error': f'SHU Result untuk {month}/{year} belum ada. Buat SHU Result terlebih dahulu.'},
            status=404,
        )

    # Ambil jasa modal pool dari shu_component_allocations (master_configuration_id=1)
    try:
        jasa_modal_alloc = ShuComponentAllocation.objects.get(
            shu_result=shu_result,
            master_configuration_id=1,
            deleted_at__isnull=True,
        )
        jasa_modal_pool = jasa_modal_alloc.allocated_amount
    except ShuComponentAllocation.DoesNotExist:
        jasa_modal_cfg = MasterConfiguration.objects.filter(
            pk=1, deleted_at__isnull=True,
        ).first()
        if not jasa_modal_cfg:
            return Response({'error': 'Konfigurasi komponen "Jasa Modal" (id=1) belum diset'}, status=400)
        jasa_modal_pool = (
            shu_result.net_profit * jasa_modal_cfg.percentage / Decimal('100')
        ).quantize(Decimal('0.01'))

    # Hitung simpanan per anggota untuk bulan ini (1 bulan)
    obligations = MemberSavingObligations.objects.filter(
        saving_type_id__in=[1, 2],
        is_active=True,
        deleted_at__isnull=True,
        member__deleted_at__isnull=True,
    ).values('member_id', 'saving_type_id', 'monthly_amount')

    member_savings = {}
    for o in obligations:
        mid = o['member_id']
        if mid not in member_savings:
            member_savings[mid] = {'wajib': Decimal('0'), 'sukarela': Decimal('0')}
        if o['saving_type_id'] == 1:
            member_savings[mid]['wajib'] += o['monthly_amount']
        elif o['saving_type_id'] == 2:
            member_savings[mid]['sukarela'] += o['monthly_amount']

    total_all_savings = sum(
        v['wajib'] + v['sukarela'] for v in member_savings.values()
    ) or Decimal('0')

    members = Members.objects.filter(deleted_at__isnull=True)
    created = updated = 0
    now_ts = timezone.now()

    try:
        with transaction.atomic():
            for m in members:
                sv = member_savings.get(m.id, {'wajib': Decimal('0'), 'sukarela': Decimal('0')})
                simp_wajib = sv['wajib']
                simp_sukarela = sv['sukarela']
                total = simp_wajib + simp_sukarela
                shu_jasa_modal = (
                    (total / total_all_savings * jasa_modal_pool).quantize(Decimal('0.01'))
                    if total_all_savings > 0 else Decimal('0')
                )
                try:
                    dist = ShuMemberDistributionsMonthly.objects.get(period=shu_result, member=m)
                    dist.simp_wajib = simp_wajib
                    dist.simp_sukarela = simp_sukarela
                    dist.total_savings = total
                    dist.total_shu = shu_jasa_modal
                    dist.status_shu = True
                    dist.updated_at = now_ts
                    dist.save(update_fields=['simp_wajib', 'simp_sukarela', 'total_savings', 'total_shu', 'status_shu', 'updated_at'])
                    updated += 1
                except ShuMemberDistributionsMonthly.DoesNotExist:
                    ShuMemberDistributionsMonthly.objects.create(
                        period=shu_result,
                        member=m,
                        simp_wajib=simp_wajib,
                        simp_sukarela=simp_sukarela,
                        total_savings=total,
                        total_shu=shu_jasa_modal,
                        status_shu=True,
                        period_month=month,
                        period_year=year,
                        created_at=now_ts,
                        updated_at=now_ts,
                    )
                    created += 1
    except Exception as exc:
        return Response(
            {'error': f'[shu_member_distributions_monthly] {type(exc).__name__}: {exc}'},
            status=500,
        )

    # Auto-sync ke tabel tahunan shu_member_distributions
    _sync_annual_distributions(year)

    MONTH_NAMES = [
        '', 'Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni',
        'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember',
    ]
    return Response({
        'message': f'SHU Jasa Modal {MONTH_NAMES[month]} {year} berhasil didistribusikan',
        'period_id': shu_result.id,
        'year': year,
        'month': month,
        'jasa_modal_pool': float(jasa_modal_pool),
        'created': created,
        'updated': updated,
        'total_members': created + updated,
    })


@api_view(['GET'])
def admin_shu_monthly_distributions(request):
    """
    GET → list shu_member_distributions_monthly untuk periode bulan tertentu.
    Auto-creates missing records for active members when ShuResult exists.
    Query params: ?year=<yyyy> &month=<1-12> &search=<nama/nik>
    """
    now = timezone.now()
    try:
        year = int(request.query_params.get('year', now.year))
        month = int(request.query_params.get('month', now.month))
    except (ValueError, TypeError):
        return Response({'error': 'year/month tidak valid'}, status=400)

    try:
        shu_result = ShuResults.objects.get(
            period_year=year, period_month=month, deleted_at__isnull=True
        )
    except ShuResults.DoesNotExist:
        return Response({'count': 0, 'results': []})

    # Resolve jasa_modal_pool
    jasa_modal_pool = None
    try:
        jasa_modal_alloc = ShuComponentAllocation.objects.get(
            shu_result=shu_result, master_configuration_id=1, deleted_at__isnull=True,
        )
        jasa_modal_pool = jasa_modal_alloc.allocated_amount
    except ShuComponentAllocation.DoesNotExist:
        cfg = MasterConfiguration.objects.filter(pk=1, deleted_at__isnull=True).first()
        if cfg:
            jasa_modal_pool = (shu_result.net_profit * cfg.percentage / Decimal('100')).quantize(Decimal('0.01'))

    # Auto-create missing member records
    if jasa_modal_pool is not None:
        obligations = MemberSavingObligations.objects.filter(
            saving_type_id__in=[1, 2], is_active=True,
            deleted_at__isnull=True, member__deleted_at__isnull=True,
        ).values('member_id', 'saving_type_id', 'monthly_amount')

        member_savings = {}
        for o in obligations:
            mid = o['member_id']
            if mid not in member_savings:
                member_savings[mid] = {'wajib': Decimal('0'), 'sukarela': Decimal('0')}
            if o['saving_type_id'] == 1:
                member_savings[mid]['wajib'] += o['monthly_amount']
            elif o['saving_type_id'] == 2:
                member_savings[mid]['sukarela'] += o['monthly_amount']

        total_all_savings = sum(
            v['wajib'] + v['sukarela'] for v in member_savings.values()
        ) or Decimal('0')

        existing_ids = set(
            ShuMemberDistributionsMonthly.objects.filter(period=shu_result)
            .values_list('member_id', flat=True)
        )

        now_ts = timezone.now()
        new_records = []
        for m in Members.objects.filter(deleted_at__isnull=True):
            if m.id not in existing_ids:
                sv = member_savings.get(m.id, {'wajib': Decimal('0'), 'sukarela': Decimal('0')})
                simp_wajib = sv['wajib']
                simp_sukarela = sv['sukarela']
                total = simp_wajib + simp_sukarela
                shu_jm = (
                    (total / total_all_savings * jasa_modal_pool).quantize(Decimal('0.01'))
                    if total_all_savings > 0 else Decimal('0')
                )
                new_records.append(ShuMemberDistributionsMonthly(
                    period=shu_result, member=m,
                    simp_wajib=simp_wajib, simp_sukarela=simp_sukarela,
                    total_savings=total, total_shu=shu_jm,
                    status_shu=True,
                    period_month=shu_result.period_month,
                    period_year=shu_result.period_year,
                    created_at=now_ts, updated_at=now_ts,
                ))
        if new_records:
            ShuMemberDistributionsMonthly.objects.bulk_create(new_records, ignore_conflicts=True)

    qs = ShuMemberDistributionsMonthly.objects.filter(
        period=shu_result
    ).select_related('member')

    search = request.query_params.get('search', '')
    if search:
        qs = qs.filter(
            Q(member__full_name__icontains=search) |
            Q(member__nik_employee__icontains=search)
        )

    data = []
    for dist in qs.order_by('member__full_name'):
        data.append({
            'id': dist.id,
            'member_id': dist.member_id,
            'member_name': dist.member.full_name,
            'simp_wajib': float(dist.simp_wajib or 0),
            'simp_sukarela': float(dist.simp_sukarela or 0),
            'total_savings': float(dist.total_savings),
            'total_shu': float(dist.total_shu),
            'distributed_status': dist.distributed_status,
            'status_shu': dist.status_shu,
        })

    return Response({'count': len(data), 'results': data})


@api_view(['PATCH', 'DELETE'])
def admin_shu_monthly_distribution_detail(request, pk):
    """
    PATCH  → edit simp_wajib/simp_sukarela (recalculates total_savings & total_shu).
    DELETE → hapus satu record distribusi bulanan.
    """
    try:
        dist = ShuMemberDistributionsMonthly.objects.select_related('member', 'period').get(pk=pk)
    except ShuMemberDistributionsMonthly.DoesNotExist:
        return Response({'error': 'Data tidak ditemukan'}, status=404)

    if request.method == 'DELETE':
        dist.delete()
        return Response(status=204)

    try:
        simp_wajib = Decimal(str(request.data.get('simp_wajib', dist.simp_wajib or 0)))
        simp_sukarela = Decimal(str(request.data.get('simp_sukarela', dist.simp_sukarela or 0)))
    except (InvalidOperation, TypeError):
        return Response({'error': 'Nilai simpanan tidak valid'}, status=400)

    total_savings = simp_wajib + simp_sukarela

    # Recalculate total_shu based on new savings
    shu_result = dist.period
    jasa_modal_pool = None
    try:
        jasa_modal_alloc = ShuComponentAllocation.objects.get(
            shu_result=shu_result, master_configuration_id=1, deleted_at__isnull=True,
        )
        jasa_modal_pool = jasa_modal_alloc.allocated_amount
    except ShuComponentAllocation.DoesNotExist:
        cfg = MasterConfiguration.objects.filter(pk=1, deleted_at__isnull=True).first()
        if cfg:
            jasa_modal_pool = (shu_result.net_profit * cfg.percentage / Decimal('100')).quantize(Decimal('0.01'))

    total_all_savings = MemberSavingObligations.objects.filter(
        saving_type_id__in=[1, 2], is_active=True,
        deleted_at__isnull=True, member__deleted_at__isnull=True,
    ).aggregate(total=Sum('monthly_amount'))['total'] or Decimal('0')

    total_shu = Decimal('0')
    if jasa_modal_pool is not None and total_all_savings > 0:
        total_shu = (total_savings / total_all_savings * jasa_modal_pool).quantize(Decimal('0.01'))

    dist.simp_wajib = simp_wajib
    dist.simp_sukarela = simp_sukarela
    dist.total_savings = total_savings
    dist.total_shu = total_shu
    dist.updated_at = timezone.now()
    dist.save(update_fields=['simp_wajib', 'simp_sukarela', 'total_savings', 'total_shu', 'updated_at'])

    # Auto-sync ke tabel tahunan
    _sync_annual_distributions(dist.period_year or shu_result.period_year)

    return Response({
        'id': dist.id,
        'member_id': dist.member_id,
        'member_name': dist.member.full_name,
        'simp_wajib': float(dist.simp_wajib),
        'simp_sukarela': float(dist.simp_sukarela),
        'total_savings': float(dist.total_savings),
        'total_shu': float(dist.total_shu),
        'distributed_status': dist.distributed_status,
        'status_shu': dist.status_shu,
    })


@api_view(['GET'])
def admin_shu_stats(request):
    """
    Ringkasan statistik SHU.
    - Jumlah periode
    - Total SHU yang sudah didistribusikan
    - Jumlah member yang sudah menerima SHU
    """
    total_periods = ShuPeriods.objects.count()
    distributed = ShuMemberDistributions.objects.filter(distributed_status=True, status_shu=True)
    return Response({
        'total_periods': total_periods,
        'total_distributed': distributed.aggregate(total=Sum('total_shu'))['total'] or 0,
        'total_recipients': distributed.values('member_id').distinct().count(),
        'latest_year': ShuPeriods.objects.order_by('-year').values_list('year', flat=True).first(),
    })


@api_view(['GET'])
def get_component_allocations(request):
    """
    GET  → list component allocations for a given period_year and period_month (13 for annual).
    """
    year_param = request.query_params.get('year')
    month_param = request.query_params.get('month')

    if not year_param:
        return Response({'error': 'year parameter is required'}, status=400)

    try:
        year = int(year_param)
        month = int(month_param) if month_param else 13
    except ValueError:
        return Response({'error': 'Invalid year or month value'}, status=400)

    try:
        shu_result = ShuResults.objects.get(period_year=year, period_month=month, deleted_at__isnull=True)
    except ShuResults.DoesNotExist:
        return Response({'error': f'SHU Result not found for period {month}/{year}'}, status=404)

    allocs = ShuComponentAllocation.objects.filter(shu_result=shu_result, deleted_at__isnull=True).order_by('id')
    
    # If no allocations exist but SHU result exists, initialize them
    if not allocs.exists():
        active_configs = MasterConfiguration.objects.filter(deleted_at__isnull=True).order_by('id')
        allocations = []
        for config in active_configs:
            allocated_amount = (shu_result.net_profit * config.percentage / Decimal('100')).quantize(Decimal('0.01'))
            allocations.append(ShuComponentAllocation(
                shu_result=shu_result,
                master_configuration=config,
                component_name=config.component_name,
                percentage=config.percentage,
                allocated_amount=allocated_amount,
                period_month=shu_result.period_month,
                period_year=shu_result.period_year,
            ))
        if allocations:
            ShuComponentAllocation.objects.bulk_create(allocations)
            allocs = ShuComponentAllocation.objects.filter(shu_result=shu_result, deleted_at__isnull=True).order_by('id')

    data = []
    for a in allocs:
        data.append({
            'id': a.id,
            'master_configuration_id': a.master_configuration_id,
            'component_name': a.component_name,
            'percentage': float(a.percentage),
            'allocated_amount': float(a.allocated_amount),
            'period_month': a.period_month,
            'period_year': a.period_year,
        })

    return Response({
        'net_profit': float(shu_result.net_profit),
        'results': data
    })


@api_view(['POST'])
def admin_shu_sync_results(request):
    """
    Recalculate and upsert shu_results for all periods (or a specific year) that have transactions.
    Called automatically on page load to ensure shu_results is always in sync.
    Body: { year?: <yyyy> }
    """
    year_param = request.data.get('year')

    qs = IncomeExpenses.objects.filter(deleted_at__isnull=True)
    if year_param:
        try:
            qs = qs.filter(transaction_date__year=int(year_param))
        except (TypeError, ValueError):
            return Response({'error': 'year tidak valid'}, status=400)

    month_dates = qs.dates('transaction_date', 'month')

    synced = 0
    for date in month_dates:
        _auto_recalculate_shu_results(date.year, date.month)
        synced += 1

    return Response({'synced': synced})


@api_view(['POST'])
def save_component_allocations(request):
    """
    POST → update component allocations for a given period and recalculate member distributions.
    Body: { year, month, allocations: [ { id, percentage }, ... ] }
    """
    year_param = request.data.get('year')
    month_param = request.data.get('month')
    allocations_data = request.data.get('allocations', [])

    if not year_param:
        return Response({'error': 'year is required'}, status=400)

    try:
        year = int(year_param)
        month = int(month_param) if month_param else 13
    except ValueError:
        return Response({'error': 'Invalid year or month value'}, status=400)

    try:
        shu_result = ShuResults.objects.get(period_year=year, period_month=month, deleted_at__isnull=True)
    except ShuResults.DoesNotExist:
        return Response({'error': 'SHU Result not found for this period'}, status=404)

    # Validate that total percentage sums to 100%
    try:
        total_pct = sum(Decimal(str(a.get('percentage', 0))) for a in allocations_data)
    except (TypeError, ValueError, InvalidOperation):
        return Response({'error': 'Nilai persentase tidak valid'}, status=400)

    if total_pct != Decimal('100'):
        return Response({'error': f'Total persentase harus 100%. Saat ini: {total_pct}%'}, status=400)

    jasa_modal_pool = Decimal('0')

    with transaction.atomic():
        for alloc_item in allocations_data:
            alloc_id = alloc_item.get('id')
            try:
                pct = Decimal(str(alloc_item.get('percentage', 0)))
            except (TypeError, ValueError, InvalidOperation):
                return Response({'error': 'Nilai persentase tidak valid'}, status=400)

            allocated_amount = (shu_result.net_profit * pct / Decimal('100')).quantize(Decimal('0.01'))

            try:
                alloc = ShuComponentAllocation.objects.get(id=alloc_id, shu_result=shu_result)
                alloc.percentage = pct
                alloc.allocated_amount = allocated_amount
                alloc.save(update_fields=['percentage', 'allocated_amount', 'updated_at'])
                
                if 'jasa modal' in alloc.component_name.lower():
                    jasa_modal_pool = allocated_amount
            except ShuComponentAllocation.DoesNotExist:
                return Response({'error': f'Allocation item with ID {alloc_id} not found for this period'}, status=404)

        # ── Recalculate Member Distributions ──
        from api.saving.models import MemberSavingObligations
        obligations = MemberSavingObligations.objects.filter(
            saving_type_id__in=[1, 2],
            is_active=True,
            deleted_at__isnull=True,
            member__deleted_at__isnull=True,
        ).values('member_id', 'saving_type_id', 'monthly_amount')

        now = timezone.now()
        if shu_result.period_month == 13:
            months_multiplier = Decimal('12') if year < now.year else Decimal(str(now.month))
        else:
            months_multiplier = Decimal('1')

        member_savings = {}
        for o in obligations:
            mid = o['member_id']
            if mid not in member_savings:
                member_savings[mid] = Decimal('0')
            member_savings[mid] += o['monthly_amount'] * months_multiplier

        total_all_savings = sum(member_savings.values()) or Decimal('0')
        members = Members.objects.filter(deleted_at__isnull=True)

        if shu_result.period_month == 13:
            # Update or create ShuMemberDistributions (annual)
            for m in members:
                total = member_savings.get(m.id, Decimal('0'))
                shu_jasa_modal = (
                    (total / total_all_savings * jasa_modal_pool).quantize(Decimal('0.01'))
                    if total_all_savings > 0 else Decimal('0')
                )
                try:
                    dist = ShuMemberDistributions.objects.get(period_year=shu_result.period_year, member=m)
                    dist.total_savings = total
                    dist.total_shu = shu_jasa_modal
                    dist.updated_at = now
                    dist.save(update_fields=['total_savings', 'total_shu', 'updated_at'])
                except ShuMemberDistributions.DoesNotExist:
                    ShuMemberDistributions.objects.create(
                        member=m,
                        total_savings=total,
                        total_shu=shu_jasa_modal,
                        period_year=shu_result.period_year,
                        created_at=now,
                        updated_at=now,
                    )
        else:
            # Update or create ShuMemberDistributionsMonthly
            for m in members:
                total = member_savings.get(m.id, Decimal('0'))
                shu_jasa_modal = (
                    (total / total_all_savings * jasa_modal_pool).quantize(Decimal('0.01'))
                    if total_all_savings > 0 else Decimal('0')
                )
                try:
                    dist = ShuMemberDistributionsMonthly.objects.get(period=shu_result, member=m)
                    dist.total_savings = total
                    dist.total_shu = shu_jasa_modal
                    dist.updated_at = now
                    dist.save(update_fields=['total_savings', 'total_shu', 'updated_at'])
                except ShuMemberDistributionsMonthly.DoesNotExist:
                    ShuMemberDistributionsMonthly.objects.create(
                        period=shu_result,
                        member=m,
                        total_savings=total,
                        total_shu=shu_jasa_modal,
                        period_month=shu_result.period_month,
                        period_year=shu_result.period_year,
                        created_at=now,
                        updated_at=now,
                    )

    return Response({'message': 'Alokasi SHU dan distribusi anggota berhasil diperbarui'})
